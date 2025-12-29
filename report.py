from customtkinter import *
import sqlite3
from tkcalendar import DateEntry
from tkinter.ttk import Style, Treeview
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from utils.type_utils import to_int, to_float
import datetime


class Report:
    def __init__(self, parent, windowControl, font="Roboto"):
        self.windowControl = windowControl
        self.font = font
        self.report = CTkToplevel(parent)
        self.report.wm_transient(parent)
        self.report.title("Report")
        self.report.geometry("700x250+350+150")
        self.report.resizable(False, False)
        self.report.protocol("WM_DELETE_WINDOW", self.destroy)

        CTkLabel(self.report, text="Report", font=(self.font, 20, "bold")).place(relx=0.5, rely=0.05, anchor=CENTER)

        self.type = StringVar(value="1")
        yesterday = datetime.date.today() - datetime.timedelta(days=1)

        CTkRadioButton(self.report, text="Transaction Date from", variable=self.type, font=(self.font, 15), value="1").place(relx=0.05, rely=0.25)
        self.fromDate = DateEntry(self.report, width=12, background='darkblue', foreground='white', font=(self.font, 10), date_pattern='dd-mm-yyyy', borderwidth=2)
        self.fromDate.place(relx=0.33, rely=0.25)
        self.fromDate.set_date(yesterday)

        CTkLabel(self.report, text="to", font=(self.font, 15)).place(relx=0.50, rely=0.25)
        self.toDate = DateEntry(self.report, width=12, background='darkblue', foreground='white', font=(self.font, 10), date_pattern='dd-mm-yyyy', borderwidth=2)
        self.toDate.place(relx=0.55, rely=0.25)

        CTkRadioButton(self.report, text="Transaction Period", variable=self.type, font=(self.font, 15), value="2").place(relx=0.05, rely=0.45)

        self.period = StringVar(value="Today")
        lists = ["Today", "Last 1 Month", "Last 2 Month", "Last 3 Month", "Last 6 Month", "Last 1 Year"]
        CTkComboBox(self.report, values=lists, state="readonly", variable=self.period, font=(self.font, 15), width=260).place(relx=0.33, rely=0.45)

        CTkButton(self.report, text="Get Data", font=(self.font, 15), width=100, command=self.getData).place(relx=0.75, rely=0.40)

        CTkLabel(self.report, text="Total Amount", font=(self.font, 15)).place(relx=0.05, rely=0.75)
        self.totalwithoutgst = DoubleVar(value=0)
        CTkEntry(self.report, width=100, state="readonly", textvariable=self.totalwithoutgst, font=(self.font, 15)).place(relx=0.20, rely=0.75)

        CTkLabel(self.report, text="Gst Amount", font=(self.font, 15)).place(relx=0.37, rely=0.75)
        self.gstamount = DoubleVar(value=0)
        CTkEntry(self.report, width=100, state="readonly", textvariable=self.gstamount, font=(self.font, 15)).place(relx=0.50, rely=0.75)

        CTkLabel(self.report, text="Net Amount", font=(self.font, 15)).place(relx=0.67, rely=0.75)
        self.netamount = DoubleVar(value=0)
        CTkEntry(self.report, width=100, state="readonly", textvariable=self.netamount, font=(self.font, 15)).place(relx=0.80, rely=0.75)

    def destroy(self):
        self.windowControl["report"] = False
        self.report.destroy()

    def findData(self, fromDate=None, toDate=None):
        # ---- RESET VALUES ----
        self.totalwithoutgst.set(0)
        self.gstamount.set(0)
        self.netamount.set(0)

        try:
            with sqlite3.connect("datas/taxinvoice.db") as conn:
                cursor = conn.cursor()

                # ---- DETERMINE DATE RANGE ----
                if self.type.get() == "2":
                    # Transaction Period mode
                    today = datetime.date.today()
                    period = self.period.get()

                    if period == "Today":
                        from_date = today
                    elif period == "Last 1 Month":
                        from_date = today - datetime.timedelta(days=30)
                    elif period == "Last 2 Month":
                        from_date = today - datetime.timedelta(days=60)
                    elif period == "Last 3 Month":
                        from_date = today - datetime.timedelta(days=90)
                    elif period == "Last 6 Month":
                        from_date = today - datetime.timedelta(days=180)
                    elif period == "Last 1 Year":
                        from_date = today - datetime.timedelta(days=365)
                    else:
                        return

                    to_date = today

                else:
                    # Transaction Date range mode
                    from_date = datetime.datetime.strptime(fromDate, "%d-%m-%Y").date()
                    to_date = datetime.datetime.strptime(toDate, "%d-%m-%Y").date()

                # ---- SQLITE SAFE DATE QUERY ----
                QUERY = """
                SELECT ti_data.quantity, ti_data.rate, ti_data.amount
                FROM ti_data
                JOIN ti_info ON ti_data.invoiceno = ti_info.invoiceno
                WHERE date(
                    substr(ti_info.invoice_date, 7, 4) || '-' ||
                    substr(ti_info.invoice_date, 4, 2) || '-' ||
                    substr(ti_info.invoice_date, 1, 2)
                ) BETWEEN date(?) AND date(?)
                """

                cursor.execute(
                    QUERY,
                    (
                        from_date.strftime("%Y-%m-%d"),
                        to_date.strftime("%Y-%m-%d")
                    )
                )

                rows = cursor.fetchall()

                # ---- CALCULATIONS ----
                for qty, rate, amount in rows:
                    qty = to_float(qty)
                    rate = to_float(rate)
                    amount = to_float(amount)

                    base = qty * rate
                    gst = amount - base

                    self.totalwithoutgst.set(
                        round(self.totalwithoutgst.get() + base, 2)
                    )
                    self.gstamount.set(
                        round(self.gstamount.get() + gst, 2)
                    )
                    self.netamount.set(
                        round(self.netamount.get() + amount, 2)
                    )

        except Exception as e:
            messagebox.showerror("ERROR", str(e))


    def getData(self):
        if self.type.get() == "1":
            if self.fromDate.get_date() > self.toDate.get_date():
                messagebox.showerror("Error", "From Date is greater than To Date")
                return
            fromDate = self.fromDate.get_date().strftime("%d-%m-%Y")
            toDate = self.toDate.get_date().strftime("%d-%m-%Y")
            self.findData(fromDate, toDate)
        else:
            self.findData(self.period.get(), datetime.date.today().strftime("%d-%m-%Y"))

class DetailedReport:
    def __init__(self, master, windowControl, font="Roboto"):
        self.master = master
        self.font = font
        self.windowControl = windowControl
        
        self.Dreport = CTkToplevel(self.master)
        self.Dreport.wm_transient(self.master) 
        self.Dreport.geometry("800x600+300+50")
        self.Dreport.title("Detailed Report")
        self.Dreport.resizable(False, False)
        self.Dreport.protocol("WM_DELETE_WINDOW", self.destroy)

        self.topFrame = CTkFrame(self.Dreport, height=150)
        self.topFrame.pack(fill="both", expand=True)
        self.bottomFrame = CTkFrame(self.Dreport)
        self.bottomFrame.pack(fill="both")

        CTkLabel(self.topFrame, text="Detailed Report", font=(self.font, 15, "bold")).place(relx=0.45, rely=0.05)
        CTkButton(self.topFrame, text="Generate Excel Report", font=(self.font, 15), command=self.generateReport).place(relx=0.78, rely=0.05)

        self.type = StringVar(value="1")
        yesterday = datetime.date.today() - datetime.timedelta(days=1)

        CTkRadioButton(self.topFrame, text="Transaction by Date from", variable=self.type, font=(self.font, 15), value="1", command=self.on_click_radio_button).place(relx=0.05, rely=0.34)
        self.fromDate = DateEntry(self.topFrame, width=12, background='darkblue', foreground='white', font=(self.font, 10), date_pattern='dd-mm-yyyy', borderwidth=2)
        self.fromDate.place(relx=0.33, rely=0.35)
        self.fromDate.set_date(yesterday)

        CTkLabel(self.topFrame, text="To", font=(self.font, 15, "bold")).place(relx=0.48, rely=0.35)
        self.toDate = DateEntry(self.topFrame, width=12, background='darkblue', foreground='white', font=(self.font, 10), date_pattern='dd-mm-yyyy', borderwidth=2)
        self.toDate.place(relx=0.52, rely=0.35)

        CTkRadioButton(self.topFrame, text="Transaction by details", variable=self.type, font=(self.font, 15), value="2", command=self.on_click_radio_button).place(relx=0.05, rely=0.65)

        values = ["BillNo", "DcNo", "Company Name", "Mobile Number"]
        self.searchBy = StringVar(value="BillNo")
        self.keyword = StringVar()
        """ CTkLabel(self.topFrame, text="SearchBy",font=(self.font, 15)).place(relx=0.05, rely=0.55) """
        self.keycbx = CTkComboBox(self.topFrame, values=values, state="readonly", variable=self.searchBy, font=(self.font, 15))
        self.keycbx.place(relx=0.30, rely=0.65)
        self.keycbx.bind("<<ComboboxSelected>>", self.on_click_combo_box)
        self.keyEnt = CTkEntry(self.topFrame, textvariable=self.keyword, font=(self.font, 15))
        self.keyEnt.place(relx=0.50, rely=0.65)

        CTkLabel(self.topFrame, text="from", font=(self.font, 15)).place(relx=0.30, rely=0.85)
        self.fromDate2 = DateEntry(self.topFrame, width=12, background='darkblue', foreground='white', font=(self.font, 10), date_pattern='dd-mm-yyyy', borderwidth=2)
        self.fromDate2.place(relx=0.35, rely=0.85)
        self.fromDate2.set_date(yesterday)

        CTkLabel(self.topFrame, text="To", font=(self.font, 15, "bold")).place(relx=0.49, rely=0.85)
        self.toDate2 = DateEntry(self.topFrame, width=12, background='darkblue', foreground='white', font=(self.font, 10), date_pattern='dd-mm-yyyy', borderwidth=2)
        self.toDate2.place(relx=0.52, rely=0.85)


        CTkButton(self.topFrame, text="Search", font=(self.font, 15), command=self.getData).place(relx=0.75, rely=0.45)

        self.on_click_combo_box()
        self.on_click_radio_button()

        tvStyle = Style()
        tvStyle.theme_use('clam')
        tvStyle.configure('Treeview', background='silver',foreground='black',rowheight=21,fieldbackground='silver')
        tvStyle.configure('mystyle.Treeview',font=(self.font,10))
        tvStyle.configure('mystyle.Treeview.Heading',font=(self.font ,10,'bold'),justify='center')

        vscrollbar = CTkScrollbar(self.bottomFrame, orientation="vertical", bg_color="black")
        vscrollbar.pack(fill="y", side="right")
        
        self.treeview = Treeview(self.bottomFrame, column=[1,2,3,4,5], height=18, show="headings", 
                                 style="mystyle.Treeview", yscrollcommand=vscrollbar.set)
        
        self.treeview.heading(1, text="Bill No")
        self.treeview.column(1, anchor="center", width=50)
        self.treeview.heading(2, text="Date")
        self.treeview.column(2, anchor="center", width=50)
        self.treeview.heading(3, text="Description")
        self.treeview.column(3, anchor="n", width=250)
        self.treeview.heading(4, text="Total Quantity")
        self.treeview.column(4, anchor="center", width=50)
        self.treeview.heading(5, text="TI/DC")
        self.treeview.column(5, anchor="center", width=100)

        self.treeview.pack(fill="x")

    def findData(self, *args):
        if args[2] == "byDate":
            QUERY1 = """SELECT invoiceno, invoice_date, buyingCompanyName, quantity FROM ti_info WHERE invoice_date BETWEEN ? AND ?;"""
            QUERY2 = """SELECT dc_id, dc_date, customer_name, quantity FROM dc_info WHERE dc_date BETWEEN ? AND ?;"""
        else:
            if args[0] == "BillNo":
                QUERY1 = """SELECT invoiceno, invoice_date, buyingCompanyName, quantity FROM ti_info WHERE invoiceno LIKE ?;"""
            if args[0] == "DcNo":
                QUERY1 = """SELECT dc_id, dc_date, buyingCompanyName, quantity FROM dc_info WHERE dc_id LIKE ?;"""
            if args[0] == "Company Name":
                QUERY1 = """SELECT invoiceno, invoice_date, buyingCompanyName, quantity FROM ti_info WHERE buyingCompanyName LIKE ?;"""
                QUERY2 = """SELECT dc_id, dc_date, buyingCompanyName, quantity FROM dc_info WHERE buyingCompanyName LIKE ?;"""
            if args[0] == "Mobile Number":
                QUERY1 = """SELECT invoiceno, invoice_date, buyingCompanyName, quantity FROM ti_info WHERE mobile_number LIKE ?;"""
                QUERY2 = """SELECT dc_id, dc_date, buyingCompanyName, quantity FROM dc_info WHERE mobile_number LIKE ?;"""

        self.conn1 = sqlite3.connect("datas/taxinvoice.db")
        self.conn2 = sqlite3.connect("datas/deliverychallan.db")

        self.cursor1 = self.conn1.cursor()
        self.cursor2 = self.conn2.cursor()

        rows = []

        if args[2] == "byDate":
            self.cursor1.execute(QUERY1, (args[0], args[1]))
            self.cursor2.execute(QUERY2, (args[0], args[1]))

            for i in self.cursor1.fetchall():
                i = list(i)
                i.append("TI")
                rows.append(i)
            for i in self.cursor2.fetchall():
                i = list(i)
                i.append("DC")
                rows.append(i)
        else:
            if args[0] == "Company Name" or args[0] == "Mobile Number":
                self.cursor1.execute(QUERY1, (args[1],))
                self.cursor2.execute(QUERY1, (args[1],))

                for i in self.cursor1.fetchall():
                    i = list(i)
                    i.append("TI")
                    rows.append(i)
                for i in self.cursor2.fetchall():
                    i = list(i)
                    i.append("DC")
                    rows.append(i)
            else:
                if args[0] == "BillNo":
                    self.cursor1.execute(QUERY1, (args[1],))

                    for i in self.cursor1.fetchall():
                        i = list(i)
                        i.append("TI")
                        rows.append(i)
                else:
                    self.cursor2.execute(QUERY1, (args[1],)) 

                    for i in self.cursor2.fetchall():
                        i = list(i)
                        i.append("DC")
                        rows.append(i)

        self.treeview.delete(*self.treeview.get_children())

        if rows == []:
            self.conn1.close()
            self.conn2.close()
            return messagebox.showerror("Error", "No Data Found")

        for i in rows:
            self.treeview.insert("", "end", values=i)

        self.conn1.close()
        self.conn2.close()

    def getData(self):
        if self.type.get() == "1":
            if self.fromDate.get_date() > self.toDate.get_date():
                messagebox.showerror("Error", "From Date is greater than To Date")
                return
        if self.type.get() == "2":
            if self.keyword.get() == "":
                messagebox.showerror("Error", "Please enter keyword")
                return

        if self.type.get() == "1":
            fromDate = self.fromDate.get_date().strftime("%d-%m-%Y")
            toDate = self.toDate.get_date().strftime("%d-%m-%Y")
            self.findData(fromDate, toDate, "byDate")
        else:
            self.findData(self.searchBy.get(), self.keyword.get(), "byData")

    def generateReport(self):
        if self.treeview.get_children() == ():
            messagebox.showerror("Error", "No Data Found")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Excel Report"

        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.append(["Bill No", "Description", "Total Quantity", "TI/DC"])

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i in self.treeview.get_children():
            values = self.treeview.item(i)["values"]
            ws.append([
                values[0],
                values[1],
                to_float(values[2]),
                values[3]
            ])


        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        try:
            wb.save("report.xlsx")
        except PermissionError:
            return messagebox.showerror("Error", "Please close report.xlsx file")
    
        self.treeview.delete(*self.treeview.get_children())

        open_file = messagebox.askyesno("Success", "Report Generated Successfully do you want to open it?")
        if open_file:
            os.startfile("report.xlsx")
        
        return
        
    def on_click_combo_box(self):
        print("happened")
        if self.searchBy.get() == "billno":
            self.fromDate2.configure(state="disabled")
            self.toDate2.configure(state="disabled")     

    def on_click_radio_button(self):
        if self.type.get() == "1":
            self.keycbx.configure(state="disabled")
            self.keyEnt.configure(state="disabled")
            self.fromDate2.configure(state="disabled")
            self.toDate2.configure(state="disabled")
            self.fromDate.configure(state="normal")
            self.toDate.configure(state="normal")
        if self.type.get() == "2":
            self.fromDate.configure(state="disabled")
            self.toDate.configure(state="disabled")
            self.keycbx.configure(state="normal")
            self.keyEnt.configure(state="normal")
            self.fromDate2.configure(state="enabled")
            self.toDate2.configure(state="enabled")

    def destroy(self):
        self.windowControl["detailed_report"] = False
        self.Dreport.destroy()
